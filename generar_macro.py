"""
generar_macro.py
================
Actualiza tablero-macro.html con datos frescos descargados de IDECBA.

Reemplaza dos bloques en el HTML:
  1. <script id="calData" ...>  → calendario de publicaciones (scrapeado)
  2. const MACRO = {...};        → datos IPCBA + PGB embebidos

Fuentes:
  - Calendario : https://www.estadisticaciudad.gob.ar/eyc/calendario-listado/
  - IPCBA      : XLSX publicado en banco_datos IDECBA
  - PGB        : XLSX publicado en banco_datos IDECBA (variación porcentual trimestral)

Dependencias:
    pip install requests openpyxl beautifulsoup4 lxml

Uso:
    python generar_macro.py
    python generar_macro.py --out tablero-macro-nuevo.html
"""

import json
import re
import sys
import argparse
from datetime import datetime, date, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl

# ── Configuración ─────────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).parent
HTML_IN      = SCRIPT_DIR / "tablero-macro.html"

IDECBA_CAL   = "https://www.estadisticaciudad.gob.ar/eyc/calendario-listado/"
CORS_PROXY   = "https://api.cors.lol/?url="          # proxy CORS usado por el tablero
IDECBA_WP    = "https://www.estadisticaciudad.gob.ar/eyc/wp-json/wp/v2/banco_datos"

IPCBA_XLSX_FALLBACK = (
    "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/"
    "2026/02/IPCBA_base_2021100-Principales_aperturas_indices.xlsx"
)
PGB_XLSX_FALLBACK = (
    "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/"
    "2025/12/PGB_K_variacion_porcentual.xlsx"
)
# Locales comerciales — ejes cuatrimestrales
LOCALES_SEARCH   = "ejes comerciales locales vacancia Ciudad Buenos Aires"
LOCALES_PATTERN  = (
    r'href="(https://www\.estadisticaciudad\.gob\.ar/eyc/wp-content/uploads/'
    r'[^"]+[Ee]jes[Cc]omerciales[^"]*\.xlsx)"'
)

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; tablero-caba/1.0)"}
TIMEOUT = 30

# ── Helpers ───────────────────────────────────────────────────────────────────

def get(url, *, proxied=False, **kw):
    target = (CORS_PROXY + requests.utils.quote(url, safe="")) if proxied else url
    r = requests.get(target, headers=HEADERS, timeout=TIMEOUT, **kw)
    r.raise_for_status()
    return r


def find_xlsx_url_via_wp(search_term, pattern):
    """Busca el URL del XLSX más reciente en el banco de datos de IDECBA."""
    params = {"search": search_term, "per_page": 5}
    try:
        posts = get(IDECBA_WP, params=params).json()
    except Exception as e:
        print(f"  ↳ wp-json no disponible ({e}), usando fallback")
        return None
    for post in posts:
        try:
            html = get(post["link"], proxied=True).text
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                return m.group(1)
        except Exception:
            continue
    return None


# ── 1. Calendario IDECBA ──────────────────────────────────────────────────────

CAT_MAP = {
    re.compile(r"\bIPCBA\b", re.I):                       ("ipcba",     "IPCBA",                      "macro"),
    re.compile(r"Canastas|l[íi]neas de pobreza", re.I):   ("canastas",  "Canastas y líneas de pobreza","macro"),
    re.compile(r"\bSIPCBA\b|precios de la construcci", re.I): ("sipcba","SIPCBA · construcción",       "macro"),
    re.compile(r"Ejes comerciales|comercio minorista", re.I):("locales", "Comercio minorista · ejes",  "macro"),
    re.compile(r"mercado laboral|ETOI", re.I):             ("empleo",    "Mercado laboral (ETOI)",      "macro"),
    re.compile(r"Ingresos en la Ciudad", re.I):            ("ingresos",  "Ingresos",                   "macro"),
    re.compile(r"Condiciones de vida", re.I):              ("condiciones","Condiciones de vida",        "macro"),
    re.compile(r"Actividad.*PGB|PGB.*trimestral", re.I):   ("iae",       "Actividad económica · PGB",  "macro"),
}

def classify_event(title):
    for pat, (cat_id, cat_label, tab) in CAT_MAP.items():
        if pat.search(title):
            return cat_id, cat_label, tab
    return "otros", "Otros", "otros"


def scrape_calendario():
    print("Descargando calendario IDECBA...")
    try:
        html = get(IDECBA_CAL).text
    except Exception as e:
        print(f"  ↳ Error scrapeando calendario: {e}")
        return None

    soup = BeautifulSoup(html, "lxml")
    hoy  = date.today()
    macro_evts, otros_evts = [], []

    for row in soup.select("tr, .calendario-row, article"):
        fecha_el  = row.select_one(".fecha, td:first-child, [class*='fecha']")
        titulo_el = row.select_one(".titulo, td:nth-child(2), h2, h3, a")
        if not fecha_el or not titulo_el:
            continue
        fecha_txt = fecha_el.get_text(strip=True)
        titulo    = titulo_el.get_text(strip=True)
        try:
            # Formatos comunes: "11/05/2026", "2026-05-11"
            if "/" in fecha_txt:
                d, m, y = fecha_txt.split("/")
                f = date(int(y), int(m), int(d))
            else:
                f = date.fromisoformat(fecha_txt[:10])
        except Exception:
            continue
        if f < hoy:
            continue

        link_el = titulo_el if titulo_el.name == "a" else titulo_el.find("a")
        url = link_el["href"] if link_el and link_el.get("href") else ""

        cat_id, cat_label, tab = classify_event(titulo)
        evt = {
            "fecha":      f.isoformat(),
            "titulo":     titulo,
            "desc_corta": titulo[:80],
            "periodo":    "",
            "cat_id":     cat_id,
            "cat_label":  cat_label,
            "tab":        tab,
            "url":        url,
        }
        (macro_evts if tab == "macro" else otros_evts).append(evt)

    # proximos por categoría
    prox = {}
    for e in macro_evts:
        if e["cat_id"] not in prox:
            prox[e["cat_id"]] = {"cat_id": e["cat_id"], "cat_label": e["cat_label"], "evento": e}

    cal = {
        "generado":              datetime.now().isoformat(timespec="seconds"),
        "fuente":                "IDECBA · " + IDECBA_CAL,
        "hoy":                   hoy.isoformat(),
        "macro":                 sorted(macro_evts, key=lambda x: x["fecha"]),
        "otros":                 sorted(otros_evts,  key=lambda x: x["fecha"]),
        "proximos_por_categoria": list(prox.values()),
        "presupuesto": [
            {"id":"pe-q1","label":"Presupuesto ejecutado · Q1 2026","descripcion":"Ejecución trimestral · estructura programática","freq":"Trimestral","ultimo_periodo":"Q4 2025","proxima_fecha":"2026-05-15","proximo_periodo":"Q1 2026","url":"https://data.buenosaires.gob.ar/dataset/presupuesto-ejecucion"},
            {"id":"pe-q2","label":"Presupuesto ejecutado · Q2 2026","descripcion":"Ejecución trimestral · estructura programática","freq":"Trimestral","ultimo_periodo":"Q1 2026","proxima_fecha":"2026-08-15","proximo_periodo":"Q2 2026","url":"https://data.buenosaires.gob.ar/dataset/presupuesto-ejecucion"},
            {"id":"cinv","label":"Cuenta de Inversión 2025","descripcion":"Ejecución anual definitiva · formato cerrado","freq":"Anual","ultimo_periodo":"2024","proxima_fecha":"2026-06-30","proximo_periodo":"2025","url":"https://www.buenosaires.gob.ar/hacienda/cuenta-de-inversion"},
            {"id":"sanc","label":"Presupuesto sancionado 2027","descripcion":"Ley anual de presupuesto","freq":"Anual","ultimo_periodo":"2026","proxima_fecha":"2026-12-15","proximo_periodo":"2027","url":"https://www.buenosaires.gob.ar/hacienda/presupuesto"},
        ],
    }
    print(f"  ✓ {len(macro_evts)} eventos macro, {len(otros_evts)} otros")
    return cal


# ── 2. IPCBA ──────────────────────────────────────────────────────────────────

MES_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def parse_ipcba_xlsx(content: bytes) -> dict:
    """Parsea el XLSX de aperturas del IPCBA (base 2021=100)."""
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

    # Busca hoja que contenga índice mensual general
    datos = {"meses": [], "indice": [], "var_mensual": [], "var_ia": [], "rubros": []}
    hoja_gral = None
    for name in wb.sheetnames:
        if re.search(r"general|gral|[íi]ndice", name, re.I):
            hoja_gral = wb[name]
            break
    if not hoja_gral:
        hoja_gral = wb.worksheets[0]

    rows = list(hoja_gral.iter_rows(values_only=True))
    # Detecta encabezado buscando columna con "Período" o "Mes"
    header_row = None
    for i, row in enumerate(rows):
        if any(str(c or "").lower() in ("período","mes","periodo","fecha") for c in row):
            header_row = i
            break
    if header_row is None:
        header_row = 0

    col_periodo = col_indice = col_var_m = col_var_ia = None
    for j, c in enumerate(rows[header_row]):
        s = str(c or "").lower()
        if "per" in s or "mes" in s or "fecha" in s:
            col_periodo = j
        elif "índice" in s or "indice" in s or "nivel" in s:
            col_indice = j
        elif "var" in s and ("m" in s or "mens" in s):
            col_var_m = j
        elif "var" in s and ("ia" in s or "anual" in s or "i.a" in s):
            col_var_ia = j

    for row in rows[header_row + 1:]:
        per = row[col_periodo] if col_periodo is not None else None
        idx = row[col_indice]  if col_indice  is not None else None
        if per is None or idx is None:
            continue
        try:
            idx = float(idx)
        except (TypeError, ValueError):
            continue

        # Formatea el período como "Ene 2024"
        if isinstance(per, datetime):
            label = f"{MES_ES[per.month - 1]} {per.year}"
        else:
            label = str(per)

        datos["meses"].append(label)
        datos["indice"].append(round(idx, 2))
        vm  = row[col_var_m]  if col_var_m  is not None else None
        via = row[col_var_ia] if col_var_ia is not None else None
        datos["var_mensual"].append(round(float(vm),  2) if vm  is not None else None)
        datos["var_ia"].append(     round(float(via), 2) if via is not None else None)

    if datos["meses"]:
        datos["ultimo_mes"]   = datos["meses"][-1]
        datos["ultimo_valor"] = datos["indice"][-1]
        datos["ultima_var_m"] = datos["var_mensual"][-1]
        datos["ultima_var_ia"]= datos["var_ia"][-1]

    # Rubros (otras hojas)
    rubros = []
    for name in wb.sheetnames:
        if name == hoja_gral.title:
            continue
        sh = wb[name]
        filas = list(sh.iter_rows(values_only=True))
        if len(filas) < 2:
            continue
        last_val = None
        last_via = None
        for row in filas[1:]:
            vals = [c for c in row if c is not None]
            if len(vals) >= 2:
                try:
                    last_val = float(vals[-2]) if len(vals) >= 2 else None
                    last_via = float(vals[-1]) if len(vals) >= 1 else None
                except (TypeError, ValueError):
                    pass
        if last_via is not None:
            rubros.append({"nombre": name.strip(), "var_ia": round(last_via, 2)})
    datos["rubros"] = sorted(rubros, key=lambda x: x["var_ia"], reverse=True)[:12]
    return datos


def fetch_ipcba() -> dict:
    print("Descargando IPCBA...")
    url = find_xlsx_url_via_wp(
        "IPCBA aperturas índice mensual",
        r'href="(https://www\.estadisticaciudad\.gob\.ar/eyc/wp-content/uploads/[^"]+IPCBA[^"]*Principales_aperturas[^"]*\.xlsx)"'
    )
    if not url:
        url = IPCBA_XLSX_FALLBACK
        print(f"  ↳ Usando URL fallback: {url}")
    else:
        print(f"  ↳ URL encontrada: {url}")

    try:
        content = get(url, proxied=True).content
        data = parse_ipcba_xlsx(content)
        print(f"  ✓ {len(data['meses'])} períodos · último: {data.get('ultimo_mes','?')}")
        return data
    except Exception as e:
        print(f"  ↳ Error descargando/parseando IPCBA: {e}")
        return {}


# ── 3. PGB ────────────────────────────────────────────────────────────────────

def parse_pgb_xlsx(content: bytes) -> dict:
    from io import BytesIO
    wb  = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws  = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # Detecta fila de encabezado (trimestres)
    header_row_idx = None
    for i, row in enumerate(rows):
        if any(re.match(r"\d{4}-T[1-4]|\d{4}T[1-4]|T[1-4]\s*\d{4}", str(c or "")) for c in row):
            header_row_idx = i
            break
    if header_row_idx is None:
        return {}

    trimestres = []
    for c in rows[header_row_idx][1:]:
        s = str(c or "").strip()
        if re.match(r"\d{4}[-\s]?T[1-4]|T[1-4][-\s]?\d{4}", s):
            trimestres.append(s.replace(" ", "-"))
        elif s:
            trimestres.append(s)

    pgb_total  = None
    sectores   = []
    for row in rows[header_row_idx + 1:]:
        nombre = str(row[0] or "").strip()
        if not nombre:
            continue
        vals = []
        for v in row[1:len(trimestres) + 1]:
            try:
                vals.append(round(float(v), 2) if v is not None else None)
            except (TypeError, ValueError):
                vals.append(None)

        if re.search(r"Producto Geogr[aá]fico Bruto|PGB total", nombre, re.I):
            pgb_total = {"valores": vals}
        elif nombre and vals:
            sectores.append({"nombre": nombre, "valores": vals})

    if not trimestres or pgb_total is None:
        return {}

    last_vals  = pgb_total["valores"]
    last_idx   = next((i for i in range(len(last_vals)-1, -1, -1) if last_vals[i] is not None), -1)
    ultimo_tri = trimestres[last_idx] if last_idx >= 0 else None

    # Var i.a. total
    var_ia = []
    for v in pgb_total["valores"]:
        var_ia.append(v)  # ya son variaciones porcentuales según el XLSX

    return {
        "trimestres":  trimestres,
        "pgb_total":   pgb_total,
        "var_ia":      var_ia,
        "sectores":    sectores[:20],
        "ultimo_trim": ultimo_tri,
        "nivel_trim":  ultimo_tri,
    }


def fetch_pgb() -> dict:
    print("Descargando PGB trimestral...")
    url = find_xlsx_url_via_wp(
        "variacion porcentual producto geografico bruto trimestral",
        r'href="(https://www\.estadisticaciudad\.gob\.ar/eyc/wp-content/uploads/[^"]+PGB[^"]*\.xlsx)"'
    )
    if not url:
        url = PGB_XLSX_FALLBACK
        print(f"  ↳ Usando URL fallback: {url}")
    else:
        print(f"  ↳ URL encontrada: {url}")

    try:
        content = get(url, proxied=True).content
        data = parse_pgb_xlsx(content)
        print(f"  ✓ {len(data.get('trimestres', []))} trimestres · último: {data.get('ultimo_trim','?')}")
        return data
    except Exception as e:
        print(f"  ↳ Error descargando/parseando PGB: {e}")
        return {}


# ── 4. Locales comerciales ────────────────────────────────────────────────────

def _periodo_label(sheet_name: str, wb_title: str = "") -> str:
    """
    Convierte el nombre de la hoja o del libro a un label de cuatrimestre.
    Formatos reconocidos: '1C2025', '2025C2', '3er cuatrimestre 2025',
    '2025-C3', 'C3 2025', etc.
    """
    s = sheet_name.strip()
    # Patron: 1C2025 / C12025 / 2025C3 / 2025-C2
    m = re.search(r'([1-3])[CcCuatrimestre]*\s*[Cc]?\s*(\d{4})|(\d{4})[-\s]*[Cc]([1-3])', s, re.I)
    if m:
        if m.group(1):
            return f"{m.group(2)}-C{m.group(1)}"
        else:
            return f"{m.group(3)}-C{m.group(4)}"
    # Patron: "1er/2do/3er cuatrimestre 2025"
    m2 = re.search(r'(\d)[ero°\. ]*(?:cuatrimestre|cuatrim|cuat).*?(\d{4})', s, re.I)
    if m2:
        return f"{m2.group(2)}-C{m2.group(1)}"
    return s


def parse_locales_xlsx(content: bytes) -> dict | None:
    """
    Parsea el XLSX de Ejes Comerciales de IDECBA.

    El XLSX puede tener:
      A) Un sheet por cuatrimestre, cada uno con filas por eje comercial.
      B) Un único sheet con columnas por período.

    Devuelve una estructura compatible con MACRO.locales:
      {
        "periodos": ["YYYY-C1", ...],  # cuatrimestres
        "ejes":     {eje: [ocu_pct, ...]},  # % ocupación paralelo a periodos
        "totales":  {eje: int},   # total locales del ÚLTIMO periodo
        "vacantes": {eje: int},   # locales vacantes del ÚLTIMO periodo
      }
    Retorna None si no puede parsear el archivo.
    """
    from io import BytesIO
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        print(f"  Error al abrir XLSX: {e}")
        return None

    # ── Detectar columnas relevantes en una hoja ──────────────────────────────
    def detect_cols(header_row):
        """
        Devuelve dict con índices de columna:
          col_eje, col_total, col_ocu_n, col_vac_n, col_ocu_pct, col_vac_pct
        """
        cols = {}
        for j, cell in enumerate(header_row):
            h = str(cell or "").lower().strip()
            if not h:
                continue
            if ("eje" in h or "denominaci" in h or "nombre" in h) and "col_eje" not in cols:
                cols["col_eje"] = j
            elif re.search(r"total|relevad|habilitad", h) and "local" in h and "col_total" not in cols:
                cols["col_total"] = j
            elif re.search(r"activ|ocup", h) and not re.search(r"tasa|%|rate", h) and "col_ocu_n" not in cols:
                cols["col_ocu_n"] = j
            elif re.search(r"vacan", h) and not re.search(r"tasa|%|rate", h) and "col_vac_n" not in cols:
                cols["col_vac_n"] = j
            elif re.search(r"tasa.*activ|tasa.*ocup|%.*activ|%.*ocup", h) and "col_ocu_pct" not in cols:
                cols["col_ocu_pct"] = j
            elif re.search(r"tasa.*vacan|%.*vacan", h) and "col_vac_pct" not in cols:
                cols["col_vac_pct"] = j
        return cols

    def find_header(rows):
        """Detecta la fila de encabezado (máx 20 filas al inicio)."""
        for i, row in enumerate(rows[:20]):
            flat = " ".join(str(c or "").lower() for c in row)
            if re.search(r"eje|local|vacan|ocup|activ", flat):
                cols = detect_cols(row)
                if "col_eje" in cols and (
                    "col_total" in cols or "col_ocu_n" in cols
                    or "col_vac_n" in cols or "col_ocu_pct" in cols
                ):
                    return i, cols
        return None, {}

    def safe_f(row, idx, default=None):
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        try:
            return round(float(v), 2) if v is not None else default
        except (TypeError, ValueError):
            return default

    def safe_i(row, idx, default=None):
        v = safe_f(row, idx)
        return int(v) if v is not None else default

    # ── Opción A: un sheet por cuatrimestre ───────────────────────────────────
    periodos = []
    ejes_ocu  = {}   # eje -> [ocu% por periodo]
    last_tot  = {}   # eje -> total (solo ultimo periodo)
    last_vac  = {}   # eje -> vacantes (solo ultimo periodo)

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hi, cols = find_header(rows)
        if hi is None or "col_eje" not in cols:
            continue

        periodo = _periodo_label(ws.title)
        if periodo not in periodos:
            periodos.append(periodo)
        pi = periodos.index(periodo)

        for row in rows[hi + 1:]:
            if not row or all(c is None for c in row):
                continue
            eje_val = row[cols["col_eje"]] if cols["col_eje"] < len(row) else None
            if not eje_val:
                continue
            eje_name = str(eje_val).strip()
            if not eje_name or re.match(r"(?i)^(total|eje|locales)", eje_name):
                continue

            ocu_pct = safe_f(row, cols.get("col_ocu_pct"))
            if ocu_pct is None:
                tot = safe_i(row, cols.get("col_total"))
                ocu_n = safe_i(row, cols.get("col_ocu_n"))
                if tot and ocu_n:
                    ocu_pct = round(ocu_n / tot * 100, 1)
            if ocu_pct is None:
                vac_pct = safe_f(row, cols.get("col_vac_pct"))
                if vac_pct is not None:
                    ocu_pct = round(100 - vac_pct, 1)

            if eje_name not in ejes_ocu:
                ejes_ocu[eje_name] = [None] * len(periodos)
            while len(ejes_ocu[eje_name]) < len(periodos):
                ejes_ocu[eje_name].append(None)
            ejes_ocu[eje_name][pi] = ocu_pct

            # Guardar conteos solo del último período parseado
            tot = safe_i(row, cols.get("col_total"))
            vac = safe_i(row, cols.get("col_vac_n"))
            if tot is not None:
                last_tot[eje_name] = tot
            if vac is not None:
                last_vac[eje_name] = vac

    if not periodos or not ejes_ocu:
        print("  No se reconoció el formato del XLSX de locales.")
        return None

    # Igualar longitudes
    n = len(periodos)
    for k in ejes_ocu:
        while len(ejes_ocu[k]) < n:
            ejes_ocu[k].append(None)

    print(f"  Locales: {len(periodos)} períodos · {len(ejes_ocu)} ejes · "
          f"{len(last_tot)} con conteos absolutos")
    return {
        "periodos": periodos,
        "ejes":     ejes_ocu,
        "totales":  last_tot,
        "vacantes": last_vac,
    }


def fetch_locales() -> dict | None:
    """Descarga y parsea el XLSX de Ejes Comerciales desde IDECBA."""
    print("Descargando Locales comerciales...")
    url = find_xlsx_url_via_wp(LOCALES_SEARCH, LOCALES_PATTERN)
    if not url:
        print("  No se encontró URL en wp-json, omitiendo locales.")
        return None
    print(f"  URL encontrada: {url}")
    try:
        content = get(url, proxied=True).content
        data = parse_locales_xlsx(content)
        if data:
            print(f"  Ultimo periodo: {data['periodos'][-1]}")
        return data
    except Exception as e:
        print(f"  Error: {e}")
        return None


# ── 5. Inyección en el HTML ───────────────────────────────────────────────────

def inyectar_cal_data(html: str, cal: dict) -> str:
    nuevo = json.dumps(cal, ensure_ascii=False, separators=(",", ":"))
    patron = r'(<script\s+id="calData"[^>]*>)([\s\S]*?)(</script>)'
    reemplazo = rf'\g<1>{nuevo}\g<3>'
    resultado, n = re.subn(patron, reemplazo, html)
    if n == 0:
        print("  ⚠ No se encontró el bloque <script id=\"calData\"> en el HTML")
    else:
        print(f"  ✓ calData inyectado ({len(nuevo):,} bytes)")
    return resultado


def extraer_macro_actual(html: str) -> dict:
    """Extrae el objeto MACRO actual del HTML para hacer merge parcial."""
    m = re.search(r'const\s+MACRO\s*=\s*(\{.*?\});\s*(?=\n|\r)', html)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    return {}


def inyectar_macro_data(html: str, macro_updates: dict) -> str:
    """
    Hace merge de macro_updates sobre el MACRO actual del HTML y lo reinyecta.
    Solo sobreescribe las claves presentes en macro_updates; el resto queda intacto.
    """
    macro_actual = extraer_macro_actual(html)
    macro_actual.update(macro_updates)
    nuevo = json.dumps(macro_actual, ensure_ascii=False, separators=(",", ":"))
    patron = r'(const\s+MACRO\s*=\s*)([\s\S]*?)(;\s*(?=\n|\r))'
    resultado, n = re.subn(patron, rf'\g<1>{nuevo}\g<3>', html, count=1)
    if n == 0:
        print("  No se encontro 'const MACRO = ...' en el HTML")
    else:
        print(f"  MACRO inyectado ({len(nuevo):,} bytes, {len(macro_actual)} claves)")
    return resultado


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Actualiza tablero-macro.html")
    parser.add_argument("--out", default=str(HTML_IN),
                        help="Archivo HTML de salida (por defecto sobreescribe el original)")
    parser.add_argument("--no-ipcba",   action="store_true", help="No actualiza IPCBA")
    parser.add_argument("--no-pgb",     action="store_true", help="No actualiza PGB")
    parser.add_argument("--no-locales", action="store_true", help="No actualiza locales comerciales")
    parser.add_argument("--no-cal",     action="store_true", help="No actualiza calendario")
    args = parser.parse_args()

    if not HTML_IN.exists():
        sys.exit(f"No se encontro: {HTML_IN}")

    html = HTML_IN.read_text(encoding="utf-8")
    print(f"Plantilla leida: {HTML_IN.name} ({len(html):,} bytes)\n")

    macro_updates = {}

    if not args.no_cal:
        cal = scrape_calendario()
        if cal:
            html = inyectar_cal_data(html, cal)

    if not args.no_ipcba:
        ipcba = fetch_ipcba()
        if ipcba:
            macro_updates["ipcba"] = ipcba

    if not args.no_pgb:
        pgb = fetch_pgb()
        if pgb:
            macro_updates["pgb"] = pgb

    if not args.no_locales:
        locales = fetch_locales()
        if locales:
            # Merge con locales existentes: extender periodos + agregar totales/vacantes
            macro_actual = extraer_macro_actual(html)
            loc_actual = macro_actual.get("locales", {})
            periodos_ant = loc_actual.get("periodos", [])
            ejes_ant     = loc_actual.get("ejes", {})

            # Incorporar periodos nuevos que no existan
            for i, per in enumerate(locales["periodos"]):
                if per not in periodos_ant:
                    periodos_ant.append(per)
                    for eje, vals in ejes_ant.items():
                        nuevo_val = locales["ejes"].get(eje, [None] * len(locales["periodos"]))
                        vals.append(nuevo_val[i] if i < len(nuevo_val) else None)
                else:
                    # Actualizar el valor del periodo existente
                    pi = periodos_ant.index(per)
                    for eje, vals in locales["ejes"].items():
                        if eje in ejes_ant:
                            ejes_ant[eje][pi] = vals[i] if i < len(vals) else None
                        else:
                            ejes_ant[eje] = [None] * len(periodos_ant)
                            ejes_ant[eje][pi] = vals[i] if i < len(vals) else None

            loc_actual["periodos"] = periodos_ant
            loc_actual["ejes"]     = ejes_ant
            loc_actual["totales"]  = locales.get("totales", {})
            loc_actual["vacantes"] = locales.get("vacantes", {})
            macro_updates["locales"] = loc_actual

    if macro_updates:
        html = inyectar_macro_data(html, macro_updates)

    out = Path(args.out)
    out.write_text(html, encoding="utf-8")
    print(f"\nGuardado en: {out} ({out.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()

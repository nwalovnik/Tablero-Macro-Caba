"""Baja PGB sectorial + IPCBA por rubro desde IDECBA, los inyecta en macro_data.json
y reemplaza el const MACRO = {...} dentro de tablero-macro.html.

Ejecutar luego de build_macro_data.py + build_calendario.py.
Pensado para el Task Scheduler de Windows o cron.
"""
import json, os, re, sys, time
import requests
from io import BytesIO
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PROJ = os.path.dirname(BASE)
MACRO_JSON = os.path.join(BASE, 'macro_data.json')
HTML_FILES = [os.path.join(BASE, 'tablero-macro.html')]
              os.path.join(BASE, 'Macro_CABA', 'tablero-macro.html'),
              os.path.join(BASE, 'Macro_CABA', 'index.html')]

H = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
     'Accept': 'application/json,text/html,*/*'}
WP_REST = 'https://www.estadisticaciudad.gob.ar/eyc/wp-json/wp/v2/banco_datos'

# ─── Helpers de descubrimiento de URL XLSX ─────────────────────────
def find_xlsx_for_search(query):
    """Busca un dataset por query en banco_datos y devuelve la URL del XLSX adjunto.
    Devuelve (None,None) si la API falla — el caller usa la URL fallback hardcoded."""
    try:
        r = requests.get(WP_REST, params={'search': query, 'per_page': 5}, headers=H, timeout=30)
        if r.status_code != 200:
            print(f'  WP REST devolvió {r.status_code}, usando fallback', flush=True)
            return None, None
        posts = r.json()
    except Exception as e:
        print(f'  WP REST falló ({e}), usando fallback', flush=True)
        return None, None
    for post in posts:
        link = post.get('link')
        if not link: continue
        try:
            html = requests.get(link, headers=H, timeout=30).text
        except Exception:
            continue
        m = re.search(r'href="(https://www\.estadisticaciudad\.gob\.ar/eyc/wp-content/uploads/[^"]+\.xlsx)"', html, re.I)
        if m: return m.group(1), post.get('title', {}).get('rendered', '')
    return None, None

def download_xlsx(url):
    r = requests.get(url, headers=H, timeout=60)
    r.raise_for_status()
    return openpyxl.load_workbook(BytesIO(r.content), data_only=True)

# ─── PGB · variación porcentual i.a. por categoría ClaNAE ─────────
def parse_pgb_variacion():
    print('[PGB var] descubriendo URL...', flush=True)
    url, title = find_xlsx_for_search('variacion porcentual producto geografico bruto trimestral')
    if not url:
        # fallback URL conocida
        url = 'https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2025/12/PGB_K_variacion_porcentual.xlsx'
        title = 'Variación porcentual i.a. del PGB Trimestral por ClaNAE (fallback)'
    print(f'[PGB var] {url}', flush=True)
    wb = download_xlsx(url)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Row 1 = header con años; Row 2 = subheader trimestres; Row 3+ = datos
    header = rows[1] or []
    sub = rows[2] or []
    year_cols = []
    for c, h in enumerate(header):
        if h is None: continue
        m = re.search(r'(20\d{2})', str(h))
        if m: year_cols.append({'year': int(m.group(1)), 'startCol': c})
    trimestres = []
    for yc in year_cols:
        for q in range(4):
            col = yc['startCol'] + q
            if col >= len(sub): continue
            s = sub[col]
            if not s: continue
            qm = re.search(r'(\d)', str(s))
            qi = int(qm.group(1)) if qm else (q+1)
            trimestres.append({'col': col, 'year': yc['year'], 'q': qi, 'label': f"{yc['year']}-T{qi}"})
    pgb_total = None
    categorias = []
    for r in rows[3:]:
        name = r[0] if r else None
        if not name: continue
        s = str(name).strip()
        if not s or s.startswith('*') or re.match(r'^Fuente', s, re.I): break
        valores = []
        for tt in trimestres:
            v = r[tt['col']] if tt['col'] < len(r) else None
            valores.append(round(float(v), 4) if isinstance(v, (int, float)) else None)
        if all(v is None for v in valores): continue
        item = {'nombre': s, 'valores': valores}
        if re.search(r'^Producto\s+Geogr', s, re.I):
            pgb_total = item
        else:
            categorias.append(item)
    if not pgb_total:
        raise RuntimeError('PGB var: no se encontró fila "Producto Geográfico Bruto"')
    # último trimestre con dato
    last_idx = len(trimestres) - 1
    while last_idx >= 0 and pgb_total['valores'][last_idx] is None:
        last_idx -= 1
    if last_idx < 0:
        raise RuntimeError('PGB var: sin valores')
    sectores_ultimo = sorted(
        [{'nombre': c['nombre'], 'var_ia': c['valores'][last_idx]} for c in categorias if c['valores'][last_idx] is not None],
        key=lambda x: -x['var_ia']
    )
    return {
        'fuente': 'IDECBA · ' + url.split('/')[-1],
        'fuente_url': url,
        'titulo_dataset': title,
        'trimestres': trimestres,
        'pgb_total': pgb_total,
        'categorias': categorias,
        'ultimo_trim': trimestres[last_idx]['label'],
        'ultimo_var_ia': pgb_total['valores'][last_idx],
        'prev_var_ia': pgb_total['valores'][last_idx-1] if last_idx > 0 else None,
        'sectores_ultimo': sectores_ultimo,
    }

# ─── PGB · nivel (precios constantes 2004) por categoría ──────────
def parse_pgb_nivel():
    """Baja la serie de PGB en millones de pesos a precios de 2004 por ClaNAE.
    Permite calcular peso de cada rama y serie de niveles."""
    print('[PGB nivel] descubriendo URL...', flush=True)
    url, title = find_xlsx_for_search('producto geografico bruto trimestral millones pesos 2004 ClaNAE')
    if not url:
        url = 'https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2025/09/PGB_K_Trimestral.xlsx'
        title = 'PGB Trimestral en millones de pesos a precios de 2004 (fallback)'
    print(f'[PGB nivel] {url}', flush=True)
    wb = download_xlsx(url)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1] or []
    sub = rows[2] or []
    year_cols = []
    for c, h in enumerate(header):
        if h is None: continue
        m = re.search(r'(20\d{2})', str(h))
        if m: year_cols.append({'year': int(m.group(1)), 'startCol': c})
    trimestres = []
    for yc in year_cols:
        for q in range(4):
            col = yc['startCol'] + q
            if col >= len(sub): continue
            s = sub[col]
            if not s: continue
            qm = re.search(r'(\d)', str(s))
            qi = int(qm.group(1)) if qm else (q+1)
            trimestres.append({'col': col, 'year': yc['year'], 'q': qi, 'label': f"{yc['year']}-T{qi}"})
    pgb_total = None
    categorias = []
    for r in rows[3:]:
        name = r[0] if r else None
        if not name: continue
        s = str(name).strip()
        if not s or s.startswith('*') or re.match(r'^Fuente', s, re.I): break
        valores = []
        for tt in trimestres:
            v = r[tt['col']] if tt['col'] < len(r) else None
            valores.append(round(float(v), 1) if isinstance(v, (int, float)) else None)
        if all(v is None for v in valores): continue
        item = {'nombre': s, 'valores': valores}
        if re.search(r'^Producto\s+Geogr', s, re.I):
            pgb_total = item
        else:
            categorias.append(item)
    return {
        'fuente': 'IDECBA · ' + url.split('/')[-1],
        'fuente_url': url,
        'trimestres': trimestres,
        'pgb_total': pgb_total,
        'categorias': categorias,
    }

# ─── IPCBA · apertura por rubro ────────────────────────────────────
DIVISIONES = [
    'Alimentos y bebidas no alcohólicas','Bebidas alcohólicas y tabaco','Prendas de vestir y calzado',
    'Vivienda, agua, electricidad, gas y otros combustibles','Equipamiento y mantenimiento del hogar',
    'Salud','Transporte','Información y comunicación','Recreación y cultura','Educación',
    'Restaurantes y hoteles','Seguros y servicios financieros',
    'Cuidado personal, protección social y otros productos'
]

def parse_ipcba_rubros():
    print('[IPCBA rubros] descubriendo URL...', flush=True)
    url, title = find_xlsx_for_search('IPCBA aperturas indice mensual')
    if not url:
        url = 'https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/02/IPCBA_base_2021100-Principales_aperturas_indices.xlsx'
        title = 'IPCBA por aperturas (fallback)'
    print(f'[IPCBA rubros] {url}', flush=True)
    wb = download_xlsx(url)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Row idx 2 = serial dates (first cell empty); row idx 3 = Nivel General
    date_row = rows[2] or []
    months = []
    for c, v in enumerate(date_row):
        if c == 0: continue
        if isinstance(v, (int, float)) and v > 30000:
            # Serial Excel a fecha (UNIX epoch 25569 = 1970-01-01)
            try:
                from datetime import datetime, timedelta
                # Excel serial date base = 1899-12-30 (with the leap-year bug)
                d = datetime(1899, 12, 30) + timedelta(days=int(v))
                months.append({'col': c, 'ym': d.strftime('%Y-%m')})
            except Exception:
                pass
        elif hasattr(v, 'strftime'):
            months.append({'col': c, 'ym': v.strftime('%Y-%m')})
    if len(months) < 13:
        raise RuntimeError(f'IPCBA rubros: solo {len(months)} meses en el XLSX')
    last_idx = len(months) - 1
    ia_idx = last_idx - 12

    def pick(name):
        for r in rows[3:]:
            cell = r[0] if r else None
            if cell and str(cell).strip() == name:
                last = r[months[last_idx]['col']] if months[last_idx]['col'] < len(r) else None
                prev = r[months[last_idx-1]['col']] if months[last_idx-1]['col'] < len(r) else None
                ia   = r[months[ia_idx]['col']] if months[ia_idx]['col'] < len(r) else None
                if not isinstance(last, (int, float)): return None
                return {
                    'nombre': name,
                    'indice': round(float(last), 2),
                    'var_mensual': round((float(last)/float(prev)-1)*100, 2) if isinstance(prev, (int, float)) else None,
                    'var_ia':      round((float(last)/float(ia)-1)*100, 2) if isinstance(ia, (int, float)) else None,
                }
        return None
    ng = pick('Nivel General')
    divisiones = [pick(n) for n in DIVISIONES]
    divisiones = [d for d in divisiones if d]
    return {
        'fuente': 'IDECBA · ' + url.split('/')[-1],
        'fuente_url': url,
        'titulo_dataset': title,
        'periodo': months[last_idx]['ym'],
        'nivel_general': ng,
        'divisiones': divisiones,
    }

# ─── Industria · ingresos fabriles por rama (re-process for "peso") ──
def industria_pesos(macro):
    """Calcula peso porcentual de cada rama sobre la suma de ramas (escala consistente)."""
    ind = macro.get('industria_ingresos', {})
    ramas = ind.get('ramas', {})
    if not ramas: return None
    last_vals = {r: (v[-1] if v and isinstance(v[-1], (int, float)) else None) for r, v in ramas.items()}
    suma = sum(v for v in last_vals.values() if v is not None)
    if not suma: return None
    pesos = {r: round(v/suma*100, 2) if v is not None else None for r, v in last_vals.items()}
    return {
        'periodo': ind.get('periodos', [None])[-1],
        'suma_ramas': round(suma, 1),
        'pesos': pesos,
    }

# ─── Patcher del HTML ──────────────────────────────────────────────
PATTERN_MACRO = re.compile(r'(const MACRO\s*=\s*)(\{.*?\})(\s*;)', re.S)

def patch_html(html_path, macro):
    if not os.path.exists(html_path):
        print(f'  (skip {html_path} — no existe)', flush=True)
        return
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    inline = json.dumps(macro, ensure_ascii=False, separators=(',', ':')).replace('</', '<\\/')
    new, n = PATTERN_MACRO.subn(lambda m: m.group(1) + inline + m.group(3), html, count=1)
    if not n:
        print(f'  WARN: no encontré const MACRO en {html_path}', flush=True)
        return
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new)
    print(f'  patched -> {os.path.basename(html_path)} ({len(inline):,} bytes inline)', flush=True)

# ─── Main ──────────────────────────────────────────────────────────
def main():
    if not os.path.exists(MACRO_JSON):
        raise RuntimeError(f'falta {MACRO_JSON}')
    with open(MACRO_JSON, 'r', encoding='utf-8') as f:
        macro = json.load(f)

    # PGB var i.a.
    pgb_var = parse_pgb_variacion()
    pgb_nivel = None
    try:
        pgb_nivel = parse_pgb_nivel()
    except Exception as e:
        print(f'[PGB nivel] falló (no es crítico): {e}', flush=True)
    # Si nivel falla, conservar pesos previos del JSON existente
    pesos_previos = None
    if not pgb_nivel and macro.get('pgb', {}).get('pesos_ultimo'):
        pesos_previos = {
            'pesos_ultimo': macro['pgb'].get('pesos_ultimo'),
            'nivel_total_ultimo': macro['pgb'].get('nivel_total_ultimo'),
            'nivel_trim': macro['pgb'].get('nivel_trim'),
        }
        print(f'[PGB nivel] conservando pesos previos ({len(pesos_previos["pesos_ultimo"])} sectores)', flush=True)

    pgb = {
        'fuente': pgb_var['fuente'],
        'titulo_dataset': pgb_var.get('titulo_dataset'),
        'trimestres': pgb_var['trimestres'],
        'pgb_total': pgb_var['pgb_total'],
        'categorias': pgb_var['categorias'],
        'ultimo_trim': pgb_var['ultimo_trim'],
        'ultimo_var_ia': pgb_var['ultimo_var_ia'],
        'prev_var_ia': pgb_var['prev_var_ia'],
        'sectores_ultimo': pgb_var['sectores_ultimo'],
    }
    if pesos_previos:
        pgb.update(pesos_previos)
    if pgb_nivel:
        # Peso sectorial sobre PGB total (último trimestre)
        last_idx = len(pgb_nivel['trimestres']) - 1
        while last_idx >= 0 and (pgb_nivel['pgb_total'] is None or pgb_nivel['pgb_total']['valores'][last_idx] is None):
            last_idx -= 1
        if last_idx >= 0 and pgb_nivel['pgb_total']:
            total_q = pgb_nivel['pgb_total']['valores'][last_idx]
            pesos = []
            for c in pgb_nivel['categorias']:
                v = c['valores'][last_idx] if last_idx < len(c['valores']) else None
                if v and total_q:
                    pesos.append({'nombre': c['nombre'], 'nivel': v, 'peso': round(v/total_q*100, 2)})
            pesos.sort(key=lambda x: -x['peso'])
            pgb['pesos_ultimo'] = pesos
            pgb['nivel_total_ultimo'] = total_q
            pgb['nivel_trim'] = pgb_nivel['trimestres'][last_idx]['label']

    macro['pgb'] = pgb
    if 'iae' in macro:
        macro['_iae_legacy'] = macro.pop('iae')
    # Sintetizo un alias `actividad` que conserva la forma esperada por el chart antiguo,
    # pero con datos del PGB: var_ia trimestral, trimestres, e índice acumulado base=100 al primer dato.
    trims_lbl = [t['label'] for t in pgb['trimestres']]
    var_ia = pgb['pgb_total']['valores']
    # Indice base 100 al primer trim usable
    idx = []
    base = None
    for v in var_ia:
        if v is None:
            idx.append(None); continue
        if base is None:
            base = 100.0
            idx.append(round(base, 2))
        else:
            # Sin un nivel real, dejamos el índice como cumulativo de variaciones (proxy)
            prev_idx = next((x for x in reversed(idx) if x is not None), None)
            if prev_idx is None:
                idx.append(round(100.0, 2))
            else:
                # No tenemos variación trimestre a trimestre, solo i.a.; dejamos índice usando i.a. como proxy
                idx.append(round(prev_idx * (1 + (v - (var_ia[var_ia.index(v)-1] if var_ia.index(v) > 0 and var_ia[var_ia.index(v)-1] is not None else 0))/100), 2))
    macro['actividad'] = {
        'trimestres': trims_lbl,
        'var_ia': var_ia,
        'fuente': 'PGB · IDECBA',
    }

    # IPCBA rubros
    rubros = parse_ipcba_rubros()
    macro.setdefault('ipcba', {})['rubros'] = rubros

    # Industria pesos
    ipesos = industria_pesos(macro)
    if ipesos:
        macro.setdefault('industria_ingresos', {})['pesos_ultimo'] = ipesos

    macro['generado_pgb_rubros'] = time.strftime('%Y-%m-%d %H:%M:%S')

    with open(MACRO_JSON, 'w', encoding='utf-8') as f:
        json.dump(macro, f, ensure_ascii=False, separators=(',', ':'))
    print(f'OK macro_data.json actualizado (PGB {len(pgb["categorias"])} cats; IPCBA rubros {len(rubros["divisiones"])}; industria pesos {len((ipesos or {}).get("pesos") or {})})', flush=True)

    # Patch HTML files
    for h in HTML_FILES:
        patch_html(h, macro)

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr)
        sys.exit(1)
